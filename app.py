#!/usr/bin/env python3
"""
================================================================================
SHARIAH COMPLIANCE SCREENER - Rule of 33%
================================================================================
Optimised for international equity screening (UK, US, Europe)
GitHub Codespaces Ready

Shariah Compliance Criteria (All must be < 33.33%):
1. Total Debt / Market Cap (or Total Assets) < 33.33%
2. Cash + Interest-Bearing Securities / Market Cap (or Total Assets) < 33.33%
3. Accounts Receivables / Market Cap (or Total Assets) < 33.33%
================================================================================
"""

import pandas as pd
import yfinance as yf
import numpy as np
from datetime import datetime
from typing import Dict, Optional, List
from concurrent.futures import ThreadPoolExecutor, as_completed
import warnings
import time
import os

warnings.filterwarnings('ignore')

# Configuration
SHARIAH_THRESHOLD = 33.33
MAX_WORKERS = 5  # Parallel threads for faster screening
RETRY_DELAY = 1  # Seconds between retries


class ShariahScreener:
    """
    Enterprise-grade Shariah Compliance Screener implementing the Rule of 33%.
    Supports international tickers across LSE, NYSE, NASDAQ, and European exchanges.
    """
    
    def __init__(self, threshold: float = SHARIAH_THRESHOLD):
        self.threshold = threshold
        self.results = None
        
    def _clean_ticker(self, ticker: str) -> Optional[str]:
        """Clean and validate ticker symbols."""
        if pd.isna(ticker) or ticker in ['UNKNOWN', '', 'nan', 'NaN']:
            return None
        
        ticker = str(ticker).strip().upper()
        
        # Handle ISIN codes (skip them - they're not valid yfinance tickers)
        if len(ticker) == 12 and ticker[:2].isalpha() and ticker[2:].isalnum():
            return None
            
        # Fix common suffix issues for yfinance compatibility
        # Handle US stocks with 'q.L' suffix (cross-listed)
        if ticker.endswith('Q.L'):
            ticker = ticker[:-3]  # Remove 'q.L' for US stocks
        
        # Handle .O and .K suffixes (NASDAQ/NYSE)
        for suffix in ['.O', '.K']:
            if ticker.endswith(suffix):
                ticker = ticker[:-2]
                break
        
        # Fix Swiss exchange suffix
        if ticker.endswith('.S') and not ticker.endswith('.AS'):
            ticker = ticker[:-2] + '.SW'
        
        # Handle spaces in tickers
        ticker = ticker.replace(' ', '')
            
        return ticker if ticker else None
    
    def _get_financial_data(self, ticker: str, retries: int = 2) -> Dict:
        for attempt in range(retries):
            try:
                stock = yf.Ticker(ticker)
                info = stock.info or {}
                
                # Check if we got valid data
                if not info or info.get('regularMarketPrice') is None:
                    if attempt < retries - 1:
                        time.sleep(RETRY_DELAY)
                        continue
                    return self._empty_result(ticker, "No data available")
                
                balance_sheet = stock.balance_sheet
                
                # Core identifiers
                company_name = info.get('longName') or info.get('shortName') or ticker
                sector = info.get('sector', 'N/A')
                industry = info.get('industry', 'N/A')
                currency = info.get('currency', 'N/A')
                
                # Market Cap
                market_cap = info.get('marketCap')
                
                # Total Assets
                total_assets = None
                if balance_sheet is not None and not balance_sheet.empty:
                    if 'Total Assets' in balance_sheet.index:
                        total_assets = balance_sheet.loc['Total Assets'].iloc[0]
                
                # Total Debt
                total_debt = info.get('totalDebt')
                if total_debt is None and balance_sheet is not None and not balance_sheet.empty:
                    long_term_debt = 0
                    short_term_debt = 0
                    
                    for col in ['Long Term Debt', 'Long Term Debt And Capital Lease Obligation']:
                        if col in balance_sheet.index:
                            val = balance_sheet.loc[col].iloc[0]
                            if val is not None and not pd.isna(val):
                                long_term_debt = val
                                break
                    
                    for col in ['Current Debt', 'Current Debt And Capital Lease Obligation', 
                               'Short Long Term Debt', 'Current Long Term Debt']:
                        if col in balance_sheet.index:
                            val = balance_sheet.loc[col].iloc[0]
                            if val is not None and not pd.isna(val):
                                short_term_debt = val
                                break
                    
                    total_debt = long_term_debt + short_term_debt
                
                # Cash & Interest-Bearing Securities
                cash = info.get('totalCash', 0) or 0
                short_term_investments = 0
                
                if balance_sheet is not None and not balance_sheet.empty:
                    if cash == 0:
                        for col in ['Cash And Cash Equivalents', 'Cash Cash Equivalents And Short Term Investments',
                                   'Cash Financial', 'Cash']:
                            if col in balance_sheet.index:
                                val = balance_sheet.loc[col].iloc[0]
                                if val is not None and not pd.isna(val):
                                    cash = val
                                    break
                    
                    for col in ['Short Term Investments', 'Other Short Term Investments',
                               'Available For Sale Securities', 'Held To Maturity Securities',
                               'Trading Securities', 'Investments And Advances']:
                        if col in balance_sheet.index:
                            val = balance_sheet.loc[col].iloc[0]
                            if val is not None and not pd.isna(val):
                                short_term_investments += val
                
                cash_and_securities = cash + short_term_investments
                
                # Accounts Receivables
                accounts_receivable = 0
                if balance_sheet is not None and not balance_sheet.empty:
                    for col in ['Net Receivables', 'Accounts Receivable', 'Receivables',
                               'Gross Accounts Receivable', 'Other Receivables']:
                        if col in balance_sheet.index:
                            val = balance_sheet.loc[col].iloc[0]
                            if val is not None and not pd.isna(val):
                                accounts_receivable = val
                                break
                
                return {
                    'ticker': ticker,
                    'company_name': company_name,
                    'sector': sector,
                    'industry': industry,
                    'currency': currency,
                    'market_cap': market_cap,
                    'total_assets': total_assets,
                    'total_debt': total_debt,
                    'cash_and_securities': cash_and_securities,
                    'accounts_receivable': accounts_receivable,
                    'data_available': True,
                    'error': None
                }
                
            except Exception as e:
                if attempt < retries - 1:
                    time.sleep(RETRY_DELAY)
                    continue
                return self._empty_result(ticker, str(e))
        
        return self._empty_result(ticker, "Max retries exceeded")
    
    def _empty_result(self, ticker: str, error: str) -> Dict:
        """Return empty result structure for failed lookups."""
        return {
            'ticker': ticker,
            'company_name': ticker,
            'sector': 'N/A',
            'industry': 'N/A',
            'currency': 'N/A',
            'market_cap': None,
            'total_assets': None,
            'total_debt': None,
            'cash_and_securities': None,
            'accounts_receivable': None,
            'data_available': False,
            'error': error
        }
    
    def _calculate_ratios(self, data: Dict) -> Dict:
        """Calculate Shariah compliance ratios."""
        result = data.copy()
        
        # Determine denominator
        denominator_mkt = data.get('market_cap')
        denominator_assets = data.get('total_assets')
        
        if denominator_mkt and denominator_mkt > 0:
            denominator = denominator_mkt
            denominator_type = 'Market Cap'
        elif denominator_assets and denominator_assets > 0:
            denominator = denominator_assets
            denominator_type = 'Total Assets'
        else:
            result['denominator_type'] = 'N/A'
            result['debt_ratio'] = None
            result['cash_ratio'] = None
            result['receivables_ratio'] = None
            result['debt_compliant'] = None
            result['cash_compliant'] = None
            result['receivables_compliant'] = None
            result['overall_compliant'] = None
            result['compliance_status'] = 'INSUFFICIENT DATA'
            return result
        
        result['denominator_type'] = denominator_type
        result['denominator_value'] = denominator
        
        # Calculate Ratios
        total_debt = data.get('total_debt') or 0
        cash_and_securities = data.get('cash_and_securities') or 0
        accounts_receivable = data.get('accounts_receivable') or 0
        
        # Debt Ratio
        debt_ratio = (total_debt / denominator) * 100
        result['debt_ratio'] = round(debt_ratio, 2)
        result['debt_compliant'] = debt_ratio < self.threshold
        
        # Cash Ratio
        cash_ratio = (cash_and_securities / denominator) * 100
        result['cash_ratio'] = round(cash_ratio, 2)
        result['cash_compliant'] = cash_ratio < self.threshold
        
        # Receivables Ratio
        receivables_ratio = (accounts_receivable / denominator) * 100
        result['receivables_ratio'] = round(receivables_ratio, 2)
        result['receivables_compliant'] = receivables_ratio < self.threshold
        
        # Overall Compliance
        if all([result['debt_compliant'], result['cash_compliant'], result['receivables_compliant']]):
            result['overall_compliant'] = True
            result['compliance_status'] = 'COMPLIANT'
        else:
            result['overall_compliant'] = False
            failed = []
            if not result['debt_compliant']:
                failed.append('Debt')
            if not result['cash_compliant']:
                failed.append('Cash')
            if not result['receivables_compliant']:
                failed.append('Recv')
            result['compliance_status'] = f'NON-COMPLIANT ({", ".join(failed)})'
        
        return result
    
    def _screen_single(self, ticker: str) -> Dict:
        """Screen a single ticker."""
        data = self._get_financial_data(ticker)
        return self._calculate_ratios(data)
    
    def screen_from_csv(self, csv_path: str, ticker_column: str = 'ticker_column',
                        parallel: bool = True) -> pd.DataFrame:
        """
        Screen tickers from CSV file with optional parallel processing.
        """
        df = pd.read_csv(csv_path)
        
        # Find ticker column
        if ticker_column not in df.columns:
            possible = ['ticker', 'tickers', 'symbol', 'ticker_column', 'Symbol', 'Ticker']
            for col in possible:
                if col in df.columns:
                    ticker_column = col
                    break
            else:
                ticker_column = df.columns[0]
        
        # Clean tickers
        raw_tickers = df[ticker_column].tolist()
        tickers = []
        skipped = 0
        
        for t in raw_tickers:
            cleaned = self._clean_ticker(t)
            if cleaned:
                tickers.append(cleaned)
            else:
                skipped += 1
        
        # Remove duplicates while preserving order
        seen = set()
        unique_tickers = []
        for t in tickers:
            if t not in seen:
                seen.add(t)
                unique_tickers.append(t)
        
        tickers = unique_tickers
        total = len(tickers)
        
        print(f"\n{'='*70}")
        print(f"  SHARIAH COMPLIANCE SCREENER - Rule of 33%")
        print(f"{'='*70}")
        print(f"  File: {csv_path}")
        print(f"  Valid tickers: {total} | Skipped (UNKNOWN/Invalid): {skipped}")
        print(f"  Threshold: < {self.threshold}% for all ratios")
        print(f"  Mode: {'Parallel' if parallel else 'Sequential'}")
        print(f"{'='*70}\n")
        
        results = []
        
        if parallel and total > 10:
            # Parallel processing for large datasets
            completed = 0
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                future_to_ticker = {executor.submit(self._screen_single, t): t for t in tickers}
                
                for future in as_completed(future_to_ticker):
                    ticker = future_to_ticker[future]
                    completed += 1
                    try:
                        result = future.result()
                        results.append(result)
                        status = result.get('compliance_status', 'ERROR')
                        print(f"  [{completed:4d}/{total}] {ticker:<12} → {status}")
                    except Exception as e:
                        results.append(self._empty_result(ticker, str(e)))
                        print(f"  [{completed:4d}/{total}] {ticker:<12} → ERROR: {str(e)[:30]}")
        else:
            # Sequential processing
            for i, ticker in enumerate(tickers, 1):
                result = self._screen_single(ticker)
                results.append(result)
                status = result.get('compliance_status', 'ERROR')
                print(f"  [{i:4d}/{total}] {ticker:<12} → {status}")
        
        # Convert to DataFrame
        results_df = pd.DataFrame(results)
        
        # Reorder columns
        column_order = [
            'ticker', 'company_name', 'sector', 'industry', 'currency',
            'compliance_status', 'overall_compliant',
            'debt_ratio', 'debt_compliant',
            'cash_ratio', 'cash_compliant',
            'receivables_ratio', 'receivables_compliant',
            'denominator_type', 'market_cap', 'total_assets',
            'total_debt', 'cash_and_securities', 'accounts_receivable',
            'data_available', 'error'
        ]
        
        existing_cols = [c for c in column_order if c in results_df.columns]
        results_df = results_df[existing_cols]
        
        self.results = results_df
        return results_df
    
    def generate_summary(self) -> str:
        """Generate summary report."""
        if self.results is None or len(self.results) == 0:
            return "No results to summarize."
        
        df = self.results
        total = len(df)
        
        compliant = len(df[df['overall_compliant'] == True])
        non_compliant = len(df[df['overall_compliant'] == False])
        insufficient_data = len(df[df['compliance_status'] == 'INSUFFICIENT DATA'])
        
        summary = f"""
{'='*70}
                    SHARIAH COMPLIANCE SCREENING SUMMARY
{'='*70}

SCREENING STATISTICS
--------------------
Total Stocks Screened:     {total:>10}
Shariah Compliant:         {compliant:>10}  ({compliant/total*100:.1f}%)
Non-Compliant:             {non_compliant:>10}  ({non_compliant/total*100:.1f}%)
Insufficient Data:         {insufficient_data:>10}  ({insufficient_data/total*100:.1f}%)

COMPLIANCE CRITERIA (Rule of 33%)
---------------------------------
All ratios must be < 33.33% to be Shariah Compliant:
1. Total Debt / Market Cap (or Total Assets)
2. Cash + Interest-Bearing Securities / Market Cap (or Total Assets)
3. Accounts Receivables / Market Cap (or Total Assets)
"""
        
        # Top compliant stocks by lowest debt ratio
        if compliant > 0:
            compliant_df = df[df['overall_compliant'] == True].sort_values('debt_ratio')
            summary += f"""
TOP 20 COMPLIANT STOCKS (by lowest debt ratio)
{'─'*50}
"""
            for _, row in compliant_df.head(20).iterrows():
                summary += f"  {row['ticker']:<12} | Debt: {row['debt_ratio']:>6.2f}% | Cash: {row['cash_ratio']:>6.2f}% | Recv: {row['receivables_ratio']:>6.2f}%\n"
        
        summary += f"""
{'='*70}
Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'='*70}
"""
        return summary
    
    def export_to_excel(self, output_path: str) -> str:
        """Export results to formatted Excel file."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        if self.results is None or len(self.results) == 0:
            return "No results to export."
        
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='1F4E79')
        compliant_fill = PatternFill('solid', fgColor='C6EFCE')
        non_compliant_fill = PatternFill('solid', fgColor='FFC7CE')
        warning_fill = PatternFill('solid', fgColor='FFEB9C')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        df = self.results
        total = len(df)
        compliant = len(df[df['overall_compliant'] == True])
        non_compliant = len(df[df['overall_compliant'] == False])
        insufficient = total - compliant - non_compliant
        
        # ─── SUMMARY SHEET ───
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        ws_summary['A1'] = 'SHARIAH COMPLIANCE SCREENING REPORT'
        ws_summary['A1'].font = Font(bold=True, size=16, color='1F4E79')
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_summary['A3'] = f'Threshold: < {self.threshold}%'
        
        stats = [
            ('Metric', 'Count', 'Percentage'),
            ('Total Screened', total, '100%'),
            ('Compliant', compliant, f'{compliant/total*100:.1f}%' if total > 0 else '0%'),
            ('Non-Compliant', non_compliant, f'{non_compliant/total*100:.1f}%' if total > 0 else '0%'),
            ('Insufficient Data', insufficient, f'{insufficient/total*100:.1f}%' if total > 0 else '0%')
        ]
        
        for i, (metric, count, pct) in enumerate(stats):
            row = 5 + i
            ws_summary[f'A{row}'] = metric
            ws_summary[f'B{row}'] = count
            ws_summary[f'C{row}'] = pct
            if i == 0:
                for col in ['A', 'B', 'C']:
                    ws_summary[f'{col}{row}'].font = header_font
                    ws_summary[f'{col}{row}'].fill = header_fill
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 15
        
        # ─── DETAILED RESULTS SHEET ───
        ws_details = wb.create_sheet('All Results')
        
        headers = ['Ticker', 'Company', 'Sector', 'Status', 
                   'Debt %', 'Cash %', 'Recv %', 'Denom Type',
                   'Market Cap', 'Total Debt', 'Cash & Sec', 'Receivables']
        
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, row in enumerate(df.itertuples(), 2):
            ws_details.cell(row=row_idx, column=1, value=row.ticker)
            ws_details.cell(row=row_idx, column=2, value=getattr(row, 'company_name', 'N/A')[:40])
            ws_details.cell(row=row_idx, column=3, value=getattr(row, 'sector', 'N/A'))
            ws_details.cell(row=row_idx, column=4, value=getattr(row, 'compliance_status', 'N/A'))
            ws_details.cell(row=row_idx, column=5, value=getattr(row, 'debt_ratio', None))
            ws_details.cell(row=row_idx, column=6, value=getattr(row, 'cash_ratio', None))
            ws_details.cell(row=row_idx, column=7, value=getattr(row, 'receivables_ratio', None))
            ws_details.cell(row=row_idx, column=8, value=getattr(row, 'denominator_type', 'N/A'))
            ws_details.cell(row=row_idx, column=9, value=getattr(row, 'market_cap', None))
            ws_details.cell(row=row_idx, column=10, value=getattr(row, 'total_debt', None))
            ws_details.cell(row=row_idx, column=11, value=getattr(row, 'cash_and_securities', None))
            ws_details.cell(row=row_idx, column=12, value=getattr(row, 'accounts_receivable', None))
            
            status = getattr(row, 'compliance_status', '')
            if 'COMPLIANT' == status:
                ws_details.cell(row=row_idx, column=4).fill = compliant_fill
            elif 'NON-COMPLIANT' in str(status):
                ws_details.cell(row=row_idx, column=4).fill = non_compliant_fill
            else:
                ws_details.cell(row=row_idx, column=4).fill = warning_fill
            
            # Highlight failed ratios
            for col_idx, attr in [(5, 'debt_ratio'), (6, 'cash_ratio'), (7, 'receivables_ratio')]:
                val = getattr(row, attr, None)
                if val is not None and val >= self.threshold:
                    ws_details.cell(row=row_idx, column=col_idx).fill = non_compliant_fill
        
        col_widths = [12, 40, 25, 30, 10, 10, 10, 12, 15, 15, 15, 15]
        for i, w in enumerate(col_widths, 1):
            ws_details.column_dimensions[chr(64+i) if i <= 26 else 'A' + chr(64+i-26)].width = w
        
        # ─── COMPLIANT ONLY SHEET ───
        ws_compliant = wb.create_sheet('Compliant Only')
        compliant_df = df[df['overall_compliant'] == True].sort_values('debt_ratio')
        
        comp_headers = ['Ticker', 'Company', 'Sector', 'Debt %', 'Cash %', 'Recv %', 'Market Cap']
        for col, header in enumerate(comp_headers, 1):
            cell = ws_compliant.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor='006400')
        
        for row_idx, row in enumerate(compliant_df.itertuples(), 2):
            ws_compliant.cell(row=row_idx, column=1, value=row.ticker)
            ws_compliant.cell(row=row_idx, column=2, value=getattr(row, 'company_name', 'N/A')[:40])
            ws_compliant.cell(row=row_idx, column=3, value=getattr(row, 'sector', 'N/A'))
            ws_compliant.cell(row=row_idx, column=4, value=getattr(row, 'debt_ratio', None))
            ws_compliant.cell(row=row_idx, column=5, value=getattr(row, 'cash_ratio', None))
            ws_compliant.cell(row=row_idx, column=6, value=getattr(row, 'receivables_ratio', None))
            ws_compliant.cell(row=row_idx, column=7, value=getattr(row, 'market_cap', None))
        
        # ─── NON-COMPLIANT SHEET ───
        ws_non = wb.create_sheet('Non-Compliant')
        non_df = df[df['overall_compliant'] == False]
        
        non_headers = ['Ticker', 'Company', 'Failed Criteria', 'Debt %', 'Cash %', 'Recv %']
        for col, header in enumerate(non_headers, 1):
            cell = ws_non.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor='8B0000')
        
        for row_idx, row in enumerate(non_df.itertuples(), 2):
            ws_non.cell(row=row_idx, column=1, value=row.ticker)
            ws_non.cell(row=row_idx, column=2, value=getattr(row, 'company_name', 'N/A')[:40])
            
            failed = []
            if getattr(row, 'debt_ratio', 0) >= self.threshold:
                failed.append('Debt')
            if getattr(row, 'cash_ratio', 0) >= self.threshold:
                failed.append('Cash')
            if getattr(row, 'receivables_ratio', 0) >= self.threshold:
                failed.append('Recv')
            ws_non.cell(row=row_idx, column=3, value=', '.join(failed))
            
            ws_non.cell(row=row_idx, column=4, value=getattr(row, 'debt_ratio', None))
            ws_non.cell(row=row_idx, column=5, value=getattr(row, 'cash_ratio', None))
            ws_non.cell(row=row_idx, column=6, value=getattr(row, 'receivables_ratio', None))
        
        wb.save(output_path)
        return output_path


def main():
    """Main entry point."""
    import sys
    
    print("""
    ╔═══════════════════════════════════════════════════════════════════════╗
    ║          SHARIAH COMPLIANCE SCREENER - Rule of 33%                    ║
    ║                                                                       ║       
    ╚═══════════════════════════════════════════════════════════════════════╝
    """)
    
    # Get CSV path
    if len(sys.argv) > 1:
        csv_path = sys.argv[1]
    else:
        csv_path = 'shariah.csv'  # Default filename
    
    if not os.path.exists(csv_path):
        print(f"  ERROR: File not found: {csv_path}")
        print("  Usage: python shariah_screener.py <path_to_csv>")
        return
    
    # Initialize and run
    screener = ShariahScreener(threshold=33.33)
    results = screener.screen_from_csv(csv_path, ticker_column='ticker_column', parallel=True)
    
    # Print summary
    print(screener.generate_summary())
    
    # Export outputs
    screener.export_to_excel('shariah_compliance_report.xlsx')
    print(f"\n  Excel report saved: shariah_compliance_report.xlsx")
    
    results.to_csv('shariah_compliance_results.csv', index=False)
    print(f"  CSV results saved: shariah_compliance_results.csv")
    
    # Quick stats
    compliant_list = results[results['overall_compliant'] == True]['ticker'].tolist()
    print(f"\n  Total Compliant Tickers: {len(compliant_list)}")
    
    return results


if __name__ == "__main__":
    main()
